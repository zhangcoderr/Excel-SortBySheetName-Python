import openpyxl
import os
import datetime
import re
import xlrd

from openpyxl import load_workbook

class SheetData:
    def __init__(self,sheetFullName,sheetKey,number,ws):
        self.sheetFullName=sheetFullName#101.表2-4措施项目工料机分析表
        self.sheetKey=sheetKey#2-4
        self.number=number#101
        self.ws_data=ws

def replace_xlsx(sheetname,sheetvalue):
    table = wb.sheet_by_name(sheetname)
    #ws2 = wb2[sheetname]
    ws2=wb2.create_sheet(sheetname)

    for i in range(table.nrows):
        for j in range(table.ncols):
            ws2.cell(row=i + 1, column=j + 1, value=table.cell(i,j).value)


    #两个for循环遍历整个excel的单元格内容
    # for i, row in enumerate(ws.iter_rows()):
    #     for j, cell in enumerate(row):
    #         ws2.cell(row=i + 1, column=j + 1, value=cell.value)

def sort_sheet():
    #name_dic={}
    #sheets=[SheetData(1,1,1,1)]
    sheets=[]
    for sheetname in sheetnames:
        numbers = re.findall('^(.*?)\.', sheetname)
        names=re.findall('表(.*?)[\u4e00-\u9fa5]',sheetname)

        ws=wb.sheet_by_name(sheetname)
        sheet=None
        if( not names):
            sheet=SheetData(sheetname,None,numbers[0],ws)
        else:

            sheet=SheetData(sheetname,names[0],numbers[0],ws)


        sheets.append(sheet)
    print(len(sheets))
    for i in range(len(sheets)):
        for j in range(len(sheets)):
            if(i==j): continue
            sheetData1 = sheets[i]
            sheetData2 = sheets[j]
            if (sheetData1 != sheetData2):
                key = sheetData1.sheetKey
                keyNext = sheetData2.sheetKey
                if (not key or not keyNext): continue
                for index, rule in enumerate(sortRule):
                    if (rule == key):
                        index_1 = index
                        break
                for index, rule in enumerate(sortRule):
                    if (rule == keyNext):
                        index_2 = index
                        break
                if (keyNext > key):
                    temp = sheetData1
                    sheets[i] = sheets[j]
                    sheets[j] = temp
    for i in range(len(sheets)):
        for j in range(len(sheets)):
            if(i==j or sheetData2.sheetKey!=sheetData1.sheetKey): continue
            sheetData1 = sheets[i]
            sheetData2 = sheets[j]
            if(sheetData1.number>sheetData2.number):
                temp = sheetData1
                sheets[i] = sheets[j]
                sheets[j] = temp
    for s in sheets:
        print(s.sheetFullName)


            #ws=sheetname#test------------------------------------

        # if name_dic.keys() and names:
        #     if sheetname in name_dic.keys():
        #         name_dic[sheetname].append(ws)
        #     else:
        #         list=[ws]
        #         name_dic.update({sheetname:list})
        # else:
        #     if names:
        #         list = [ws]
        #         name_dic.update({sheetname: list})
        #     else:
        #         list = [ws]
        #         name_dic.update({sheetname: list})
    # for sheetname1 in name_dic:                 #排序有问题 todotodotodotodoooooooooooooooo
    #     for sheetname2 in name_dic:
    #         numbers = re.findall('^(.*?)\.', sheetname)
    #
    #         keys=re.findall('表(.*?)[\u4e00-\u9fa5]',sheetname1)
    #         keyNexts=re.findall('表(.*?)[\u4e00-\u9fa5]',sheetname2)
    #
    #         if(not keys or not keyNexts):
    #             continue
    #         key=keys[0]
    #         keyNext=keyNexts[0]
    #
    #         index_1=-1
    #         index_2=-1
    #         for index,rule in enumerate(sortRule):
    #             if(rule==key):
    #                 index_1=index
    #                 break
    #         for index,rule in enumerate(sortRule):
    #             if(rule==keyNext):
    #                 index_2=index
    #                 break
    #
    #         if(index_1!=-1 and index_2!=-1):
    #             if(index_2>index_1):
    #                 temp=name_dic.pop(sheetname1)
    #                 name_dic[sheetname1]=name_dic.pop(sheetname2)
    #                 name_dic[sheetname2]=temp
    #
    #                 temp=name_dic[sheetname1]
    #                 name_dic[sheetname1]=name_dic[sheetname2]
    #                 name_dic[sheetname2]=temp       #排序有问题 todotodotodotodooooooooo

    # for sheetname1 in name_dic:
    #     for sheetArray in name_dic[sheetname1]:
    #         if(len(name_dic[sheetname1])==1):
    #             break
    #         for sheetArray2 in name_dic[sheetname1]:
    #             number1=re.findall('^(.*?)\.',sheetname1)[0]
    #             number2=re.findall('^(.*?)\.',sheetname2)[0]
    #             if(number2<number1):
    #                 temp = name_dic.pop(sheetname1)
    #                 name_dic[sheetname1] = name_dic.pop(sheetname2)
    #                 name_dic[sheetname2] = temp
    #
    #                 temp = name_dic[sheetname1]
    #                 name_dic[sheetname1] = name_dic[sheetname2]
    #                 name_dic[sheetname2] = temp


    return sheets
if __name__ == "__main__":
    os.chdir(r"C:\Users\123\Desktop\ExcelSort")
    sortRule=['1-1', '1-1-1', '1-1-2', '1-2','1-3-A', '1-3-B','1-3-C',  '1-3-A-1', '1-4', '1-4-1', '1-4-2', '1-5', '1-6', '1-7',  '2-1','2-2', '2-3', '2-4','2-5', '2-7',]

    filename = 'all.xlsx'
    filename2 = 'testResult.xlsx'
    print('loading')

    wb=xlrd.open_workbook(filename)
    print('load done')
    wb2 = load_workbook(filename2)


    sheetnames = wb.sheet_names()
    result= sort_sheet()
    print(result)

    # for sheetname in dic_result:
    #     print(sheetname)
    #     print(dic_result[sheetname])
    #     replace_xlsx(sheetname,dic_result[sheetname])
    # for sheetname in sheetnames:
    #     replace_xlsx(sheetname)


    # d= datetime.datetime.now().strftime('%d-%M-%S')
    # filename2='testResult'+d+'.xlsx'
    # wb2.save(filename2)
